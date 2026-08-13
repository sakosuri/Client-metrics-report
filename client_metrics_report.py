#!/usr/bin/env python3
"""
Fetch per-client metrics from Cisco Catalyst Center and export averaged results to Excel.

APIs used:
  POST /dna/data/api/v1/clients/{id}/trendAnalytics  -> RSSI, SNR, Onboarding/Roaming times, Tx/Rx rates
  GET  /dna/data/api/v1/clients/{id}                 -> Health score, connected device MAC/name

Input:  Excel file with client MAC addresses in the first column (header row optional).
Output: Excel with one row per client, columns:
        MAC Address | RSSI | SNR | Onboarding Time | Roaming Time |
        Tx Rate | Rx Rate | Health Score | Connected Device MAC | Connected Device Name

Each metric is the average of all 5-minute max values across the selected time window.

Environment variables (.env in this folder):
  CATC_API_BASE   e.g. https://<catc--ip>
  CATC_AUTH_URL   e.g. https://<catc-ip>/dna/system/api/v1/auth/token
  CATC_USERNAME
  CATC_PASSWORD

Usage examples:
  python client_metrics_report.py --time-range 24h --input clients.xlsx
  python client_metrics_report.py --time-range 7d  --input clients.xlsx --output report.xlsx
"""

import argparse
import datetime as dt
import io
import os
import statistics
import time
from pathlib import Path

import matplotlib
matplotlib.use("Agg")  # non-interactive backend; must precede pyplot import
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
import requests
import urllib3
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

REQUEST_TIMEOUT = 60
DEFAULT_PAGE_LIMIT = 500

# Attribute names as accepted by the per-client trendAnalytics endpoint.
# Adjust these if the CatC version uses different names (check API explorer).
TREND_ATTRIBUTES = [
    {"name": "rssi",               "function": "max"},
    {"name": "snr",                "function": "max"},
    {"name": "txRate",             "function": "max"},
    {"name": "rxRate",             "function": "max"},
    {"name": "dataRate",           "function": "max"},
    {"name": "maxRunDuration",     "function": "max"},
    {"name": "maxRoamingDuration", "function": "max"},
]

# Header row values that should be skipped when reading the MAC input file
_MAC_COLUMN_HEADERS = {"mac address", "mac", "client mac", "macaddress"}


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

def load_config():
    env_path = Path(__file__).parent / ".env"
    if env_path.exists():
        load_dotenv(str(env_path))

    api_base  = os.environ.get("CATC_API_BASE",  "").strip()
    auth_url  = os.environ.get("CATC_AUTH_URL",  "").strip()
    username  = os.environ.get("CATC_USERNAME",  "").strip()
    password  = os.environ.get("CATC_PASSWORD",  "").strip()

    missing = [n for n, v in [
        ("CATC_API_BASE", api_base), ("CATC_AUTH_URL", auth_url),
        ("CATC_USERNAME", username), ("CATC_PASSWORD", password),
    ] if not v]
    if missing:
        raise ValueError(f"Missing required environment values: {', '.join(missing)}")

    return api_base.rstrip("/"), auth_url, username, password


# ---------------------------------------------------------------------------
# Authentication
# ---------------------------------------------------------------------------

def obtain_token(auth_url: str, username: str, password: str) -> str:
    resp = requests.post(
        auth_url,
        auth=(username, password),
        headers={"Content-Type": "application/json"},
        timeout=REQUEST_TIMEOUT,
        verify=False,
    )
    resp.raise_for_status()
    token = resp.json().get("Token")
    if not token:
        raise RuntimeError("Token not found in auth response.")
    return token


# ---------------------------------------------------------------------------
# Time range helpers
# ---------------------------------------------------------------------------

def get_time_range(selection: str):
    """Return (start_epoch_ms, end_epoch_ms) based on user selection."""
    now_ms = int(time.time() * 1000)
    offsets = {"24h": 24 * 3600 * 1000, "7d": 7 * 24 * 3600 * 1000}
    if selection not in offsets:
        raise ValueError(f"Unknown time range '{selection}'. Choose '24h' or '7d'.")
    return now_ms - offsets[selection], now_ms


# ---------------------------------------------------------------------------
# Input: read MAC addresses from Excel
# ---------------------------------------------------------------------------

def read_mac_addresses(excel_path: str) -> list[str]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    macs = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        val = row[0] if row else None
        if not val:
            continue
        mac = str(val).strip()
        if mac.lower() in _MAC_COLUMN_HEADERS:
            continue
        macs.append(mac)
    wb.close()
    return macs


# ---------------------------------------------------------------------------
# API calls
# ---------------------------------------------------------------------------

def _post_with_pagination(endpoint: str, headers: dict, base_body: dict) -> list:
    all_items = []
    cursor = None
    seen_cursors: set[str] = set()

    while True:
        body = {**base_body, "page": {"limit": DEFAULT_PAGE_LIMIT, "timeSortOrder": "asc"}}
        if cursor:
            body["page"]["cursor"] = cursor

        resp = requests.post(endpoint, headers=headers, json=body,
                             timeout=REQUEST_TIMEOUT, verify=False)

        if resp.status_code == 404:
            break
        if resp.status_code >= 400:
            try:
                detail = resp.json()
            except Exception:
                detail = resp.text
            print(f"    API warning [{resp.status_code}]: {detail}")
            break

        payload = resp.json()
        all_items.extend(payload.get("response", []))

        next_cursor = payload.get("page", {}).get("cursor")
        if not next_cursor or next_cursor in seen_cursors:
            break
        seen_cursors.add(next_cursor)
        cursor = next_cursor

    return all_items


def fetch_client_trend(api_base: str, token: str, mac: str,
                       start_ms: int, end_ms: int) -> list:
    endpoint = f"{api_base}/dna/data/api/v1/clients/{mac}/trendAnalytics"
    headers  = {"X-Auth-Token": token, "Content-Type": "application/json"}
    body = {
        "startTime":           start_ms,
        "endTime":             end_ms,
        "trendInterval":       "5MIN",
        "aggregateAttributes": TREND_ATTRIBUTES,
    }
    return _post_with_pagination(endpoint, headers, body)


def fetch_client_details(api_base: str, token: str, mac: str, at_time_ms: int) -> dict:
    """Fetch point-in-time client info for health score and connected device."""
    endpoint = f"{api_base}/dna/intent/api/v1/client-detail"
    headers  = {"X-Auth-Token": token, "Content-Type": "application/json"}
    params   = {"macAddress": mac, "timestamp": at_time_ms}

    resp = requests.get(endpoint, headers=headers, params=params,
                        timeout=REQUEST_TIMEOUT, verify=False)
    if resp.status_code == 404:
        return {}
    if resp.status_code >= 400:
        try:
            err = resp.json()
        except Exception:
            err = resp.text
        print(f"    Details API warning [{resp.status_code}]: {err}")
        return {}

    return resp.json().get("detail", {}) or {}


# ---------------------------------------------------------------------------
# Metric extraction helpers
# ---------------------------------------------------------------------------

def _collect_values(trend_items: list, attr_name: str) -> list[float]:
    """Collect every non-None max value for attr_name across all trend items."""
    name_lower = attr_name.lower()
    values: list[float] = []

    for item in trend_items:
        for agg in item.get("aggregateAttributes") or []:
            if agg.get("name", "").lower() == name_lower:
                v = agg.get("value")
                if v is not None:
                    try:
                        values.append(float(v))
                    except (TypeError, ValueError):
                        pass
        for group in item.get("groups") or []:
            for agg in group.get("aggregateAttributes") or []:
                if agg.get("name", "").lower() == name_lower:
                    v = agg.get("value")
                    if v is not None:
                        try:
                            values.append(float(v))
                        except (TypeError, ValueError):
                            pass
    return values


def average_of_maxes(trend_items: list, attr_name: str):
    """Return the mean of all 5-min max samples, or None if no data."""
    values = _collect_values(trend_items, attr_name)
    if not values:
        return None
    return round(statistics.mean(values), 4)


_IST = dt.timezone(dt.timedelta(hours=5, minutes=30))


def _to_ist(epoch_ms) -> str:
    try:
        return dt.datetime.fromtimestamp(int(epoch_ms) / 1000, tz=_IST).strftime("%Y-%m-%d %H:%M:%S IST")
    except Exception:
        return ""


def _ms_to_s(value):
    """Milliseconds to seconds, rounded to 3 dp."""
    if value is None:
        return None
    return round(float(value) / 1000, 3)


def _bps_to_kbps(value):
    """Bytes per second to kilobytes per second, rounded to 3 dp."""
    if value is None:
        return None
    return round(float(value) / 1000, 3)


def _bits_to_kbps(value):
    """Bits per second to kilobits per second (÷ 1000), rounded to 3 dp."""
    if value is None:
        return None
    return round(float(value) / 1000, 3)


def extract_interval_rows(mac: str, trend_items: list) -> list:
    """Return one dict per 5-min interval with unit-converted values."""
    attr_map = {
        "rssi":             "rssi",
        "snr":              "snr",
        "maxrunduration":   "onboarding_duration_s",
        "maxroamingduration": "roaming_duration_s",
        "txrate":           "tx_rate_kbps",
        "rxrate":           "rx_rate_kbps",
    }
    interval_rows = []
    for item in trend_items:
        ts = item.get("timestamp")
        row = {"mac": mac, "timestamp_ist": _to_ist(ts) if ts else ""}

        def _pick(attr_name_lower):
            for agg in item.get("aggregateAttributes") or []:
                if agg.get("name", "").lower() == attr_name_lower:
                    return agg.get("value")
            for group in item.get("groups") or []:
                for agg in group.get("aggregateAttributes") or []:
                    if agg.get("name", "").lower() == attr_name_lower:
                        return agg.get("value")
            return None

        raw_rssi  = _pick("rssi")
        raw_snr   = _pick("snr")
        raw_onb   = _pick("maxrunduration")
        raw_roam  = _pick("maxroamingduration")
        raw_tx    = _pick("txrate")
        raw_rx    = _pick("rxrate")
        raw_dr    = _pick("datarate")

        row["rssi"]                  = float(raw_rssi) if raw_rssi is not None else None
        row["snr"]                   = float(raw_snr)  if raw_snr  is not None else None
        row["onboarding_duration_s"] = _ms_to_s(raw_onb)
        row["roaming_duration_s"]    = _ms_to_s(raw_roam)
        row["tx_rate_kbps"]          = _bps_to_kbps(raw_tx)
        row["rx_rate_kbps"]          = _bps_to_kbps(raw_rx)
        row["data_rate"]             = _bits_to_kbps(float(raw_dr) if raw_dr is not None else None)
        interval_rows.append(row)
    return interval_rows


def fetch_assurance_events(api_base: str, token: str, mac: str,
                           start_ms: int, end_ms: int) -> list:
    endpoint = f"{api_base}/dna/data/api/v1/assuranceEvents"
    headers  = {"X-Auth-Token": token, "Content-Type": "application/json"}
    all_events = []
    offset = 1
    limit  = 20

    while True:
        params = {
            "deviceFamily": "Wireless Client",
            "clientMac":    mac,
            "startTime":    start_ms,
            "endTime":      end_ms,
            "limit":        limit,
            "offset":       offset,
        }
        resp = requests.get(endpoint, headers=headers, params=params,
                            timeout=REQUEST_TIMEOUT, verify=False)
        if resp.status_code == 404:
            break
        if resp.status_code >= 400:
            try:
                err = resp.json()
            except Exception:
                err = resp.text
            print(f"    Events API warning [{resp.status_code}]: {err}")
            break

        payload = resp.json()
        page = payload.get("response", []) or []
        all_events.extend(page)
        if len(page) < limit:
            break
        offset += limit

    return all_events


def extract_full_detail_metrics(details: dict) -> dict:
    """Extract all point-in-time metrics from a client-detail response."""
    health_score = None
    for entry in details.get("healthScore") or []:
        if isinstance(entry, dict) and entry.get("healthType", "").upper() == "OVERALL":
            health_score = entry.get("score")
            break
    else:
        hl = details.get("healthScore") or []
        if hl and isinstance(hl[0], dict):
            health_score = hl[0].get("score")

    connected_devices = details.get("connectedDevice") or []
    connected_mac  = connected_devices[0].get("mac",  "") if connected_devices else ""
    connected_name = connected_devices[0].get("name", "") if connected_devices else ""

    def _safe_float(v):
        try:
            return float(v) if v is not None else None
        except (TypeError, ValueError):
            return None

    onboarding = details.get("onboarding") or {}

    return {
        "rssi":               _safe_float(details.get("rssi")),
        "snr":                _safe_float(details.get("snr")),
        "tx_rate_kbps":       _bps_to_kbps(_safe_float(details.get("txRate"))),
        "rx_rate_kbps":       _bps_to_kbps(_safe_float(details.get("rxRate"))),
        "data_rate":          _bits_to_kbps(_safe_float(details.get("dataRate"))),
        "health_score":       health_score,
        "connected_mac":      connected_mac,
        "connected_name":     connected_name,
        # ms -> s for durations sourced from onboarding sub-object
        "onboarding_duration_s": _ms_to_s(_safe_float(onboarding.get("maxRunDuration"))),
        "roaming_duration_s":    _ms_to_s(_safe_float(details.get("maxRoamingDuration"))),
    }


def build_event_rows(api_base: str, token: str, mac: str, events: list) -> list:
    """For each event, fetch client-detail at event time and assemble a report row."""
    rows = []
    for event in events:
        ts_ms      = event.get("timestamp") or event.get("eventTimestamp")
        event_name = (event.get("name") or event.get("eventType")
                      or event.get("reasonDescription") or "")
        event_lower = event_name.lower()
        is_onboarding = any(k in event_lower for k in ("onboard", "assoc", "auth", "dhcp"))
        is_roaming    = "roam" in event_lower

        details = fetch_client_details(api_base, token, mac, ts_ms) if ts_ms else {}
        m = extract_full_detail_metrics(details)

        rows.append({
            "mac":                   mac,
            "event_name":            event_name,
            "event_timestamp_ist":   _to_ist(ts_ms) if ts_ms else "",
            "rssi":                  m["rssi"],
            "snr":                   m["snr"],
            "tx_rate_kbps":          m["tx_rate_kbps"],
            "rx_rate_kbps":          m["rx_rate_kbps"],
            "data_rate":             m["data_rate"],
            "health_score":          m["health_score"],
            "connected_mac":         m["connected_mac"],
            "connected_name":        m["connected_name"],
            "onboarding_duration_s": m["onboarding_duration_s"] if is_onboarding else None,
            "roaming_duration_s":    m["roaming_duration_s"]    if is_roaming    else None,
        })
    return rows


def extract_client_info(details: dict):
    """Pull health score and connected device info from the client details response."""
    health_score = None
    health_list = details.get("healthScore") or []
    # Prefer OVERALL score; fall back to the first entry
    for entry in health_list:
        if isinstance(entry, dict) and entry.get("healthType", "").upper() == "OVERALL":
            health_score = entry.get("score")
            break
    else:
        if health_list and isinstance(health_list[0], dict):
            health_score = health_list[0].get("score")

    connected_mac  = ""
    connected_name = ""
    connected_devices = details.get("connectedDevice") or []
    if connected_devices and isinstance(connected_devices[0], dict):
        connected_mac  = connected_devices[0].get("mac",  "")
        connected_name = connected_devices[0].get("name", "")

    return health_score, connected_mac, connected_name


# ---------------------------------------------------------------------------
# Chart helpers
# ---------------------------------------------------------------------------

_OBD_LABELS  = ["<2s", "2-4s", "4-6s", "6-8s", "8-10s", ">=10s", "fail"]
_ROAM_LABELS = ["<=300ms", "301-500ms", "501-1000ms", "1001-3000ms", ">3000ms", "fail"]
_RSSI_LABELS = [">=-45", "-55 to -45", "-65 to -55", "-71 to -65", "<=-72"]
_SNR_LABELS  = [">=40", "20 to 40", "15 to 20", "10 to 15", "<10"]


def _bucket_onboarding(val_s):
    if val_s is None: return 6
    if val_s < 2:     return 0
    if val_s < 4:     return 1
    if val_s < 6:     return 2
    if val_s < 8:     return 3
    if val_s < 10:    return 4
    return 5


def _bucket_roaming(val_s):
    if val_s is None: return 5
    ms = val_s * 1000
    if ms <= 300:  return 0
    if ms <= 500:  return 1
    if ms <= 1000: return 2
    if ms <= 3000: return 3
    return 4


def _bucket_rssi(val):
    if val is None:  return None
    if val >= -45:   return 0
    if val >= -55:   return 1
    if val >= -65:   return 2
    if val >= -71:   return 3
    return 4


def _bucket_snr(val):
    if val is None: return None
    if val >= 40:   return 0
    if val >= 20:   return 1
    if val >= 15:   return 2
    if val >= 10:   return 3
    return 4


def _make_bar_chart(labels, counts, title, xlabel,
                    threshold_idx=None, threshold_label=None,
                    summary_text=None) -> io.BytesIO:
    # Extra top margin gives room for the summary text above the axes
    fig, ax = plt.subplots(figsize=(5.5, 4.2))
    fig.subplots_adjust(top=0.78, bottom=0.22, left=0.12, right=0.97)

    ax.bar(range(len(labels)), counts, color="#5a3e8b", width=0.22, zorder=3)

    if threshold_idx is not None:
        ax.axvspan(threshold_idx - 0.5, len(labels) - 0.5,
                   alpha=0.12, color="red", zorder=0)

    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=20, ha="right", fontsize=7)
    ax.set_ylabel("Total Clients", fontsize=8)
    ax.set_xlabel(xlabel, fontsize=8)
    ax.set_title(title, fontsize=10, fontweight="bold", pad=4)
    ax.yaxis.get_major_locator().set_params(integer=True)
    ax.grid(axis="y", linestyle="--", linewidth=0.4, alpha=0.5, zorder=0)

    # Place both labels in the figure area above the axes, outside the plot
    if summary_text:
        fig.text(0.02, 0.95, summary_text,
                 fontsize=10, color="green", va="top", ha="left")
    if threshold_label:
        fig.text(0.97, 0.95,
                 f"▪ {threshold_label}",
                 fontsize=10, color="#c0392b", va="top", ha="right")

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=110, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf


def generate_chart_images(summary_rows: list) -> list:
    """Return [(BytesIO, title), ...] for onboarding, roaming, RSSI, SNR charts."""
    obd_counts  = [0] * len(_OBD_LABELS)
    roam_counts = [0] * len(_ROAM_LABELS)
    rssi_counts = [0] * len(_RSSI_LABELS)
    snr_counts  = [0] * len(_SNR_LABELS)

    for row in summary_rows:
        obd_counts[_bucket_onboarding(row.get("onboarding_time"))] += 1
        roam_counts[_bucket_roaming(row.get("roaming_time"))]       += 1
        r = _bucket_rssi(row.get("rssi"))
        if r is not None: rssi_counts[r] += 1
        s = _bucket_snr(row.get("snr"))
        if s is not None: snr_counts[s]  += 1

    n = len(summary_rows) or 1
    pct_obd  = sum(obd_counts[:5]) / n * 100           # < 10 s
    pct_roam = sum(roam_counts[:4]) / n * 100           # <= 3000 ms
    rt = sum(rssi_counts) or 1
    st = sum(snr_counts)  or 1
    pct_rssi = sum(rssi_counts[:4]) / rt * 100          # > -72 dBm
    pct_snr  = sum(snr_counts[:4])  / st * 100          # >= 10 dB

    return [
        (
            _make_bar_chart(
                _OBD_LABELS, obd_counts,
                "Client Onboarding Times", "Time (Seconds)",
                threshold_idx=5, threshold_label=">= 10s Threshold",
                summary_text=f"{pct_obd:.0f}% clients with onboarding times < 10 s",
            ),
            "Client Onboarding Times",
        ),
        (
            _make_bar_chart(
                _ROAM_LABELS, roam_counts,
                "Client Roaming Times", "Time (ms)",
                threshold_idx=4, threshold_label="Over Threshold",
                summary_text=f"{pct_roam:.0f}% clients with roaming times < 3000 ms",
            ),
            "Client Roaming Times",
        ),
        (
            _make_bar_chart(
                _RSSI_LABELS, rssi_counts,
                "Connectivity RSSI", "RSSI (dBm)",
                threshold_idx=4, threshold_label="<= -72 dBm Threshold",
                summary_text=f"{pct_rssi:.0f}% clients with RSSI > -72 dBm",
            ),
            "Connectivity RSSI",
        ),
        (
            _make_bar_chart(
                _SNR_LABELS, snr_counts,
                "Connectivity SNR", "SNR (dB)",
                threshold_idx=4, threshold_label="< 10 dB Threshold",
                summary_text=f"{pct_snr:.0f}% clients with SNR > 10 dB",
            ),
            "Connectivity SNR",
        ),
    ]


# ---------------------------------------------------------------------------
# Output: write Excel report
# ---------------------------------------------------------------------------

def write_output_excel(summary_rows: list, interval_rows: list,
                       event_rows: list, output_path: str):
    wb = Workbook()

    # --- Sheet 1: summary (one row per client) ---
    ws_summary = wb.active
    ws_summary.title = "Client Metrics Summary"
    ws_summary.append([
        "MAC Address",
        "RSSI (avg max, dBm)",
        "SNR (avg max, dB)",
        "Onboarding Duration (avg max, s)",
        "Roaming Duration (avg max, s)",
        "Tx Rate (avg max, kBps)",
        "Rx Rate (avg max, kBps)",
        "Data Rate (avg max, kbps)",
        "Health Score",
        "Connected Device MAC",
        "Connected Device Name",
    ])
    for row in summary_rows:
        ws_summary.append([
            row["mac"],
            row["rssi"],
            row["snr"],
            row["onboarding_time"],
            row["roaming_time"],
            row["tx_rate"],
            row["rx_rate"],
            row["data_rate"],
            row["health_score"],
            row["connected_mac"],
            row["connected_name"],
        ])

    # --- Sheet 2: raw 5-min intervals ---
    ws_raw = wb.create_sheet(title="5-Min Interval Data")
    ws_raw.append([
        "MAC Address",
        "Timestamp (IST)",
        "RSSI (dBm)",
        "SNR (dB)",
        "Onboarding Duration (s)",
        "Roaming Duration (s)",
        "Rx Rate (kBps)",
        "Tx Rate (kBps)",
        "Data Rate (kbps)",
    ])
    for row in interval_rows:
        ws_raw.append([
            row["mac"],
            row["timestamp_ist"],
            row["rssi"],
            row["snr"],
            row["onboarding_duration_s"],
            row["roaming_duration_s"],
            row["rx_rate_kbps"],
            row["tx_rate_kbps"],
            row["data_rate"],
        ])

    # --- Sheet 3: events ---
    ws_events = wb.create_sheet(title="Client Events")
    ws_events.append([
        "MAC Address",
        "Event",
        "Event Timestamp (IST)",
        "RSSI (dBm)",
        "SNR (dB)",
        "Tx Rate (kBps)",
        "Rx Rate (kBps)",
        "Data Rate (kbps)",
        "Health Score",
        "Connected Device MAC",
        "Connected Device Name",
        "Onboarding Duration (s)",
        "Roaming Duration (s)",
    ])
    for row in event_rows:
        ws_events.append([
            row["mac"],
            row["event_name"],
            row["event_timestamp_ist"],
            row["rssi"],
            row["snr"],
            row["tx_rate_kbps"],
            row["rx_rate_kbps"],
            row["data_rate"],
            row["health_score"],
            row["connected_mac"],
            row["connected_name"],
            row["onboarding_duration_s"],
            row["roaming_duration_s"],
        ])

    # --- Sheet 4: charts (2×2 grid) ---
    ws_charts = wb.create_sheet(title="Charts")
    # anchor positions: (onboarding, roaming) top row; (RSSI, SNR) bottom row
    anchors = ["A1", "L1", "A26", "L26"]
    for (buf, _title), anchor in zip(generate_chart_images(summary_rows), anchors):
        img = XLImage(buf)
        ws_charts.add_image(img, anchor)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description="Fetch per-client CatC metrics and export averaged results to Excel.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python client_metrics_report.py --time-range 24h --input clients.xlsx\n"
            "  python client_metrics_report.py --time-range 7d  --input clients.xlsx --output report.xlsx\n"
        ),
    )
    parser.add_argument(
        "--time-range", choices=["24h", "7d"],
        help="'24h' for the last 24 hours, '7d' for the last 7 days (prompted if omitted)",
    )
    parser.add_argument(
        "--input", default="client_metrics_report_input.xlsx",
        help="Excel (.xlsx) file with client MAC addresses in the first column",
    )
    today = dt.datetime.now().strftime("%d_%m_%y")
    parser.add_argument(
        "--output", default=f"client_metrics_report_{today}.xlsx",
        help="Output Excel filename (default: client_metrics_report_DD_MM_YY.xlsx)",
    )
    return parser.parse_args()


def prompt_time_range() -> str:
    print("\nSelect time range:")
    print("  1. Last 24 hours")
    print("  2. Last 7 days")
    while True:
        choice = input("Enter 1 or 2: ").strip()
        if choice == "1":
            return "24h"
        if choice == "2":
            return "7d"
        print("  Invalid choice. Please enter 1 or 2.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = parse_args()
    time_range = args.time_range or prompt_time_range()
    api_base, auth_url, username, password = load_config()

    print("Authenticating...")
    token = obtain_token(auth_url, username, password)
    print("  Token obtained.")

    start_ms, end_ms = get_time_range(time_range)
    print(f"Time range: {time_range}  [{start_ms} -> {end_ms}]")

    mac_list = read_mac_addresses(args.input)
    print(f"Clients to process: {len(mac_list)}")

    summary_rows  = []
    interval_rows = []
    event_rows    = []
    for idx, mac in enumerate(mac_list, 1):
        print(f"\n[{idx}/{len(mac_list)}] {mac}")

        trend_items = fetch_client_trend(api_base, token, mac, start_ms, end_ms)
        print(f"  Trend data points: {len(trend_items)}")

        # Point-in-time details use current time (when report is executed)
        details = fetch_client_details(api_base, token, mac, end_ms)
        health_score, connected_mac, connected_name = extract_client_info(details)

        raw_onb_avg  = average_of_maxes(trend_items, "maxRunDuration")
        raw_roam_avg = average_of_maxes(trend_items, "maxRoamingDuration")
        raw_tx_avg   = average_of_maxes(trend_items, "txRate")
        raw_rx_avg   = average_of_maxes(trend_items, "rxRate")
        raw_dr_avg   = average_of_maxes(trend_items, "dataRate")

        summary_rows.append({
            "mac":             mac,
            "rssi":            average_of_maxes(trend_items, "rssi"),
            "snr":             average_of_maxes(trend_items, "snr"),
            "onboarding_time": _ms_to_s(raw_onb_avg),
            "roaming_time":    _ms_to_s(raw_roam_avg),
            "tx_rate":         _bps_to_kbps(raw_tx_avg),
            "rx_rate":         _bps_to_kbps(raw_rx_avg),
            "data_rate":       _bits_to_kbps(raw_dr_avg),
            "health_score":    health_score,
            "connected_mac":   connected_mac,
            "connected_name":  connected_name,
        })
        interval_rows.extend(extract_interval_rows(mac, trend_items))

        events = fetch_assurance_events(api_base, token, mac, start_ms, end_ms)
        print(f"  Events fetched: {len(events)}")
        event_rows.extend(build_event_rows(api_base, token, mac, events))

    write_output_excel(summary_rows, interval_rows, event_rows, args.output)
    print(f"\nReport saved: {args.output}")
    print(f"Clients written: {len(summary_rows)}, interval rows: {len(interval_rows)}, event rows: {len(event_rows)}")


if __name__ == "__main__":
    main()
